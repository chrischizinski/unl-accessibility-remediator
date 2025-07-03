"""
Excel Spreadsheet (.xlsx) Accessibility Processor for UNL Accessibility Remediator

This module analyzes Excel spreadsheets for WCAG 2.1 Level AA compliance issues
and provides recommendations for improvement. Essential for university data accessibility.
"""

import os
import json
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any
import logging

# Excel processing
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as e:
    logging.warning(f"Excel processing libraries not available: {e}")
    openpyxl = None

# Text analysis
try:
    import textstat
except ImportError:
    textstat = None

from .contrast_checker import ContrastChecker

logger = logging.getLogger(__name__)

class XlsxAccessibilityProcessor:
    """Analyzes Excel spreadsheets (.xlsx) for accessibility compliance"""
    
    def __init__(self):
        self.contrast_checker = ContrastChecker()
        self.issues = []
        self.fixes_applied = []
        
    def analyze_xlsx(self, file_path: str, apply_fixes: bool = False) -> Dict[str, Any]:
        """
        Analyze an Excel spreadsheet for accessibility issues
        
        Args:
            file_path: Path to the .xlsx file
            apply_fixes: Whether to attempt automatic fixes
            
        Returns:
            Dictionary containing analysis results
        """
        logger.info(f"Starting Excel spreadsheet accessibility analysis: {file_path}")
        
        if not openpyxl:
            return self._create_error_result("openpyxl library not available")
        
        try:
            # Reset for new analysis
            self.issues = []
            self.fixes_applied = []
            
            # Basic file validation
            if not os.path.exists(file_path):
                return self._create_error_result(f"File not found: {file_path}")
            
            # Open and analyze Excel workbook
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            workbook_info = self._extract_workbook_info(workbook)
            worksheets = self._analyze_worksheets(workbook)
            
            # Perform accessibility checks
            self._check_workbook_structure(workbook_info, worksheets)
            self._check_worksheet_accessibility(worksheets)
            self._check_data_accessibility(worksheets)
            self._check_formatting_accessibility(worksheets)
            
            # Calculate accessibility score
            score = self._calculate_accessibility_score()
            
            # Apply fixes if requested
            if apply_fixes:
                output_path = self._apply_automatic_fixes(file_path, workbook)
            else:
                output_path = None
            
            workbook.close()
            
            return self._create_analysis_result(
                file_path=file_path,
                output_path=output_path,
                score=score,
                workbook_info=workbook_info,
                worksheets=worksheets
            )
            
        except Exception as e:
            logger.error(f"Error analyzing Excel spreadsheet: {str(e)}")
            return self._create_error_result(f"Analysis failed: {str(e)}")
    
    def _extract_workbook_info(self, workbook) -> Dict[str, Any]:
        """Extract basic workbook metadata and properties"""
        info = {
            "title": None,
            "author": None,
            "subject": None,
            "description": None,
            "worksheets": len(workbook.worksheets),
            "worksheet_names": [ws.title for ws in workbook.worksheets],
            "has_hidden_sheets": False,
            "has_protected_sheets": False
        }
        
        try:
            # Core properties
            props = workbook.properties
            if props:
                info["title"] = props.title
                info["author"] = props.creator
                info["subject"] = props.subject
                info["description"] = props.description
            
            # Check for hidden or protected sheets
            for ws in workbook.worksheets:
                if ws.sheet_state == 'hidden':
                    info["has_hidden_sheets"] = True
                if ws.protection.sheet:
                    info["has_protected_sheets"] = True
                    
        except Exception as e:
            logger.warning(f"Error extracting workbook info: {str(e)}")
        
        return info
    
    def _analyze_worksheets(self, workbook) -> List[Dict[str, Any]]:
        """Analyze all worksheets in the workbook"""
        worksheets = []
        
        for ws_idx, worksheet in enumerate(workbook.worksheets):
            try:
                ws_info = {
                    "index": ws_idx,
                    "name": worksheet.title,
                    "is_hidden": worksheet.sheet_state == 'hidden',
                    "is_protected": worksheet.protection.sheet,
                    "has_data": False,
                    "data_range": None,
                    "tables": [],
                    "headers": [],
                    "merged_cells": [],
                    "formulas": [],
                    "charts": [],
                    "hyperlinks": [],
                    "cell_count": 0,
                    "color_usage": []
                }
                
                # Analyze data range
                if worksheet.max_row > 1 or worksheet.max_column > 1:
                    ws_info["has_data"] = True
                    ws_info["data_range"] = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
                    ws_info["cell_count"] = worksheet.max_row * worksheet.max_column
                
                # Check for tables
                for table in worksheet.tables.values():
                    ws_info["tables"].append({
                        "name": table.name,
                        "range": str(table.ref),
                        "has_headers": table.headerRowCount > 0,
                        "style": table.tableStyleInfo.name if table.tableStyleInfo else None
                    })
                
                # Analyze headers (assume first row/column are headers)
                if ws_info["has_data"]:
                    # First row as headers
                    first_row = []
                    for col in range(1, min(worksheet.max_column + 1, 20)):  # Limit to 20 columns
                        cell = worksheet.cell(1, col)
                        if cell.value:
                            first_row.append({
                                "column": col,
                                "value": str(cell.value),
                                "is_bold": cell.font.bold if cell.font else False,
                                "has_fill": bool(cell.fill.start_color.rgb != 'FFFFFFFF' if cell.fill else False)
                            })
                    if first_row:
                        ws_info["headers"].append({"type": "row", "data": first_row})
                    
                    # First column as headers
                    first_col = []
                    for row in range(1, min(worksheet.max_row + 1, 100)):  # Limit to 100 rows
                        cell = worksheet.cell(row, 1)
                        if cell.value:
                            first_col.append({
                                "row": row,
                                "value": str(cell.value),
                                "is_bold": cell.font.bold if cell.font else False,
                                "has_fill": bool(cell.fill.start_color.rgb != 'FFFFFFFF' if cell.fill else False)
                            })
                    if first_col and len(first_col) > 1:  # More than just the header
                        ws_info["headers"].append({"type": "column", "data": first_col[:10]})  # First 10 rows
                
                # Check merged cells
                for merged_range in worksheet.merged_cells.ranges:
                    ws_info["merged_cells"].append(str(merged_range))
                
                # Sample formulas
                formula_count = 0
                for row in worksheet.iter_rows(max_row=min(50, worksheet.max_row), max_col=min(20, worksheet.max_column)):
                    for cell in row:
                        if cell.data_type == 'f' and formula_count < 10:  # Sample first 10 formulas
                            ws_info["formulas"].append({
                                "cell": cell.coordinate,
                                "formula": cell.value
                            })
                            formula_count += 1
                
                # Check for charts
                ws_info["charts"] = [{
                    "type": type(chart).__name__,
                    "title": chart.title.tx.rich.p[0].r.t if chart.title and chart.title.tx and chart.title.tx.rich else "No title"
                } for chart in worksheet._charts]
                
                # Sample hyperlinks
                for hyperlink in worksheet._hyperlinks[:10]:  # Sample first 10
                    ws_info["hyperlinks"].append({
                        "cell": hyperlink.ref,
                        "target": hyperlink.target,
                        "display": hyperlink.display
                    })
                
                # Sample color usage
                color_samples = []
                sample_count = 0
                for row in worksheet.iter_rows(max_row=min(20, worksheet.max_row), max_col=min(10, worksheet.max_column)):
                    for cell in row:
                        if sample_count >= 50:  # Limit samples
                            break
                        if cell.font and cell.font.color and cell.font.color.rgb:
                            if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                                color_samples.append({
                                    "cell": cell.coordinate,
                                    "font_color": cell.font.color.rgb,
                                    "fill_color": cell.fill.start_color.rgb,
                                    "value": str(cell.value)[:50] if cell.value else ""
                                })
                                sample_count += 1
                
                ws_info["color_usage"] = color_samples
                worksheets.append(ws_info)
                
            except Exception as e:
                logger.warning(f"Error analyzing worksheet '{worksheet.title}': {str(e)}")
                # Add minimal worksheet info on error
                worksheets.append({
                    "index": ws_idx,
                    "name": worksheet.title,
                    "error": str(e),
                    "has_data": False
                })
        
        return worksheets
    
    def _check_workbook_structure(self, workbook_info: Dict, worksheets: List[Dict]):
        """Check workbook structure accessibility"""
        
        # Check for workbook title
        if not workbook_info.get("title"):
            self.issues.append({
                "type": "structure",
                "severity": "medium",
                "issue": "Missing workbook title",
                "description": "Workbook should have a descriptive title in properties",
                "recommendation": "Add a descriptive title in File > Info > Properties",
                "wcag_criterion": "2.4.2 Page Titled"
            })
        
        # Check for meaningful worksheet names
        generic_names = ["Sheet1", "Sheet2", "Sheet3", "Worksheet"]
        for ws in worksheets:
            if ws["name"] in generic_names and ws.get("has_data", False):
                self.issues.append({
                    "type": "structure",
                    "severity": "medium",
                    "issue": f"Generic worksheet name: '{ws['name']}'",
                    "description": "Worksheets should have descriptive names",
                    "recommendation": f"Rename '{ws['name']}' to describe its content",
                    "wcag_criterion": "2.4.6 Headings and Labels"
                })
        
        # Check for hidden worksheets
        if workbook_info.get("has_hidden_sheets"):
            self.issues.append({
                "type": "structure",
                "severity": "low",
                "issue": "Contains hidden worksheets",
                "description": "Hidden worksheets may contain inaccessible content",
                "recommendation": "Review hidden worksheets for important content",
                "wcag_criterion": "4.1.2 Name, Role, Value"
            })
    
    def _check_worksheet_accessibility(self, worksheets: List[Dict]):
        """Check individual worksheet accessibility"""
        
        for ws in worksheets:
            if not ws.get("has_data", False):
                continue
            
            ws_name = ws["name"]
            
            # Check for proper table structure
            if not ws["tables"] and ws.get("cell_count", 0) > 50:
                self.issues.append({
                    "type": "tables",
                    "severity": "medium",
                    "issue": f"'{ws_name}': Large data range not structured as table",
                    "description": "Large data sets should use Excel tables for accessibility",
                    "recommendation": "Convert data range to Excel table (Insert > Table)",
                    "wcag_criterion": "1.3.1 Info and Relationships"
                })
            
            # Check for merged cells (problematic for screen readers)
            if ws["merged_cells"]:
                self.issues.append({
                    "type": "tables",
                    "severity": "medium",
                    "issue": f"'{ws_name}': Contains merged cells ({len(ws['merged_cells'])} ranges)",
                    "description": "Merged cells can be problematic for screen readers",
                    "recommendation": "Avoid merged cells; use cell formatting instead",
                    "wcag_criterion": "1.3.2 Meaningful Sequence",
                    "details": ws["merged_cells"][:5]  # Show first 5
                })
            
            # Check for proper headers
            if ws.get("cell_count", 0) > 20 and not ws["headers"]:
                self.issues.append({
                    "type": "tables",
                    "severity": "high",
                    "issue": f"'{ws_name}': No clear header structure detected",
                    "description": "Data tables should have clear row or column headers",
                    "recommendation": "Use bold formatting or table headers for first row/column",
                    "wcag_criterion": "1.3.1 Info and Relationships"
                })
    
    def _check_data_accessibility(self, worksheets: List[Dict]):
        """Check data content accessibility"""
        
        for ws in worksheets:
            if not ws.get("has_data", False):
                continue
            
            ws_name = ws["name"]
            
            # Check hyperlinks
            for link in ws.get("hyperlinks", []):
                if not link.get("display") or link["display"] in ["Click here", "Link", "More info"]:
                    self.issues.append({
                        "type": "links",
                        "severity": "medium",
                        "issue": f"'{ws_name}': Vague hyperlink text in {link['cell']}",
                        "description": "Hyperlink text should be descriptive",
                        "recommendation": "Use descriptive text that explains the link destination",
                        "wcag_criterion": "2.4.4 Link Purpose"
                    })
            
            # Check for formulas without clear purpose
            if len(ws.get("formulas", [])) > 5:
                self.issues.append({
                    "type": "content",
                    "severity": "low",
                    "issue": f"'{ws_name}': Many formulas may need documentation",
                    "description": "Complex formulas should be documented for accessibility",
                    "recommendation": "Add comments to explain complex calculations",
                    "wcag_criterion": "3.2.4 Consistent Identification"
                })
    
    def _check_formatting_accessibility(self, worksheets: List[Dict]):
        """Check formatting and color accessibility"""
        
        for ws in worksheets:
            ws_name = ws["name"]
            
            # Check color usage
            color_only_formatting = 0
            for color_sample in ws.get("color_usage", []):
                # Simple check for color-only differentiation
                if color_sample["font_color"] != "FF000000":  # Not default black
                    color_only_formatting += 1
            
            if color_only_formatting > 5:
                self.issues.append({
                    "type": "color",
                    "severity": "medium", 
                    "issue": f"'{ws_name}': Extensive use of color formatting",
                    "description": "Information should not be conveyed by color alone",
                    "recommendation": "Use symbols, patterns, or text in addition to color",
                    "wcag_criterion": "1.4.1 Use of Color"
                })
            
            # Check for color contrast (simplified)
            for color_sample in ws.get("color_usage", []):
                if color_sample.get("font_color") and color_sample.get("fill_color"):
                    try:
                        # Basic contrast check (simplified - would need full color analysis)
                        font_color = color_sample["font_color"]
                        fill_color = color_sample["fill_color"]
                        
                        # Flag potential contrast issues (simplified heuristic)
                        if font_color == fill_color or (font_color.upper() == "FFFFFFFF" and fill_color.upper() == "FFFFFFFF"):
                            self.issues.append({
                                "type": "color",
                                "severity": "high",
                                "issue": f"'{ws_name}': Potential contrast issue in {color_sample['cell']}",
                                "description": "Text and background colors may not have sufficient contrast",
                                "recommendation": "Verify color contrast meets WCAG AA standards (4.5:1 ratio)",
                                "wcag_criterion": "1.4.3 Contrast (Minimum)"
                            })
                            break  # Only report once per worksheet
                    except Exception:
                        pass  # Skip color analysis on error
    
    def _calculate_accessibility_score(self) -> int:
        """Calculate overall accessibility score"""
        if not self.issues:
            return 100
        
        severity_weights = {
            "critical": 20,
            "high": 10,
            "medium": 5,
            "low": 2
        }
        
        total_deductions = sum(
            severity_weights.get(issue["severity"], 5) 
            for issue in self.issues
        )
        
        score = max(0, 100 - total_deductions)
        return score
    
    def _apply_automatic_fixes(self, file_path: str, workbook) -> str:
        """Apply automatic fixes to the spreadsheet"""
        
        try:
            # Create output path
            input_path = Path(file_path)
            output_path = input_path.parent / "output" / f"accessible_{input_path.name}"
            output_path.parent.mkdir(exist_ok=True)
            
            fixes_count = 0
            
            # Fix 1: Add workbook title if missing
            if not workbook.properties.title:
                title = input_path.stem.replace('_', ' ').title()
                workbook.properties.title = title
                fixes_count += 1
                self.fixes_applied.append({
                    "type": "metadata",
                    "description": f"Added workbook title: {title}"
                })
            
            # Fix 2: Rename generic worksheet names
            generic_names = ["Sheet1", "Sheet2", "Sheet3"]
            for idx, ws in enumerate(workbook.worksheets):
                if ws.title in generic_names and ws.max_row > 1:
                    new_name = f"Data_{idx + 1}"
                    ws.title = new_name
                    fixes_count += 1
                    self.fixes_applied.append({
                        "type": "structure",
                        "description": f"Renamed generic sheet to '{new_name}'"
                    })
            
            # Fix 3: Format potential headers with bold
            for ws in workbook.worksheets:
                if ws.max_row > 1 and ws.max_column > 1:
                    # Make first row bold (potential headers)
                    for col in range(1, min(ws.max_column + 1, 10)):
                        cell = ws.cell(1, col)
                        if cell.value and not (cell.font and cell.font.bold):
                            cell.font = Font(bold=True)
                            fixes_count += 1
                    
                    if fixes_count > 0:
                        self.fixes_applied.append({
                            "type": "formatting",
                            "description": f"Applied bold formatting to headers in '{ws.title}'"
                        })
            
            # Save the modified workbook
            workbook.save(str(output_path))
            
            logger.info(f"Applied {fixes_count} automatic fixes to {output_path}")
            return str(output_path)
            
        except Exception as e:
            logger.error(f"Error applying fixes: {str(e)}")
            self.fixes_applied.append({
                "type": "error",
                "description": f"Could not apply automatic fixes: {str(e)}"
            })
            return None
    
    def _create_analysis_result(self, file_path: str, score: int, output_path: str = None, **kwargs) -> Dict[str, Any]:
        """Create standardized analysis result"""
        return {
            "success": True,
            "file_path": file_path,
            "output_path": output_path,
            "file_type": "xlsx",
            "accessibility_score": score,
            "total_issues": len(self.issues),
            "critical_issues": len([i for i in self.issues if i["severity"] == "critical"]),
            "high_issues": len([i for i in self.issues if i["severity"] == "high"]),
            "medium_issues": len([i for i in self.issues if i["severity"] == "medium"]),
            "low_issues": len([i for i in self.issues if i["severity"] == "low"]),
            "issues": self.issues,
            "fixes_applied": self.fixes_applied,
            "workbook_info": kwargs.get("workbook_info", {}),
            "worksheets": kwargs.get("worksheets", []),
            "recommendations": self._generate_recommendations(),
            "wcag_compliance": self._assess_wcag_compliance()
        }
    
    def _create_error_result(self, error_message: str) -> Dict[str, Any]:
        """Create error result"""
        return {
            "success": False,
            "error": error_message,
            "file_type": "xlsx",
            "accessibility_score": 0,
            "total_issues": 0,
            "issues": [],
            "fixes_applied": []
        }
    
    def _generate_recommendations(self) -> List[str]:
        """Generate prioritized recommendations"""
        recommendations = []
        
        critical_issues = [i for i in self.issues if i["severity"] == "critical"]
        high_issues = [i for i in self.issues if i["severity"] == "high"]
        
        if critical_issues:
            recommendations.append("❗ CRITICAL: Address table structure and data accessibility")
            
        if high_issues:
            recommendations.append("🔥 HIGH PRIORITY: Add proper headers and improve table structure")
            
        recommendations.extend([
            "📊 Convert data ranges to Excel tables for better structure",
            "🏷️ Use descriptive names for worksheets and workbook",
            "🎨 Ensure color is not the only way to convey information",
            "🔗 Use descriptive text for hyperlinks",
            "✅ Test with screen readers after making improvements"
        ])
        
        return recommendations
    
    def _assess_wcag_compliance(self) -> Dict[str, str]:
        """Assess WCAG 2.1 Level AA compliance"""
        critical_issues = [i for i in self.issues if i["severity"] == "critical"]
        high_issues = [i for i in self.issues if i["severity"] == "high"]
        
        if critical_issues:
            return {
                "level": "Non-compliant",
                "status": "Major accessibility barriers present",
                "next_steps": "Address critical table structure issues before using in courses"
            }
        elif high_issues:
            return {
                "level": "Partially compliant",
                "status": "Some accessibility improvements needed",
                "next_steps": "Address high-priority issues for better data accessibility"
            }
        else:
            return {
                "level": "Mostly compliant",
                "status": "Meets most WCAG 2.1 AA requirements for data tables",
                "next_steps": "Address remaining minor issues for full compliance"
            }